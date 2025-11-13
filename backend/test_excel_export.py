"""
Test Suite for Excel Export Module
Tests Excel export functionality including formatting, color-coding, and tips inclusion.
"""

import pytest
import os
import sys
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Add backend directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from excel_exporter import (
    ExcelExporter,
    export_jobs_to_excel,
    export_jobs_to_file
)


# Sample test data
SAMPLE_JOBS = [
    {
        'title': 'Senior Python Developer',
        'company': 'Tech Corp',
        'location': 'San Francisco, CA',
        'salary': '$120,000 - $150,000',
        'job_type': 'Remote',
        'description': 'Looking for an experienced Python developer with strong backend skills...',
        'link': 'https://example.com/job1',
        'score': {
            'overall_score': 85,
            'highlight': 'green',
            'component_scores': {
                'keyword_match': 90,
                'salary_match': 80,
                'location_match': 75,
                'job_type_match': 95
            }
        }
    },
    {
        'title': 'Junior Developer',
        'company': 'Startup Inc',
        'location': 'New York, NY',
        'salary': '$60,000 - $80,000',
        'job_type': 'Onsite',
        'description': 'Entry-level position for new graduates with basic programming knowledge...',
        'link': 'https://example.com/job2',
        'score': {
            'overall_score': 45,
            'highlight': 'yellow',
            'component_scores': {
                'keyword_match': 40,
                'salary_match': 50,
                'location_match': 45,
                'job_type_match': 45
            }
        }
    },
    {
        'title': 'Data Scientist',
        'company': 'Analytics Ltd',
        'location': 'Austin, TX',
        'salary': '$90,000 - $110,000',
        'job_type': 'Hybrid',
        'description': 'Looking for data scientist with ML experience...',
        'link': 'https://example.com/job3',
        'score': {
            'overall_score': 30,
            'highlight': 'red',
            'component_scores': {
                'keyword_match': 25,
                'salary_match': 35,
                'location_match': 30,
                'job_type_match': 30
            }
        }
    }
]

SAMPLE_TIPS = {
    'summary': 'Your resume shows good technical skills but needs improvement in quantifying achievements and adding more relevant keywords.',
    'overall_assessment': {
        'strength_score': 72,
        'completeness': 'Good',
        'ats_compatibility': 'Fair'
    },
    'critical_tips': [
        {
            'category': 'keywords',
            'title': 'Add Missing Technical Keywords',
            'description': 'Your resume is missing important keywords like Python, AWS, Docker that appear frequently in target jobs',
            'action': 'Add these keywords naturally in your experience section, especially in project descriptions',
            'impact': 'high'
        },
        {
            'category': 'formatting',
            'title': 'Improve ATS Compatibility',
            'description': 'Your resume format may not be ATS-friendly',
            'action': 'Use standard section headers and avoid tables or complex formatting',
            'impact': 'high'
        }
    ],
    'important_tips': [
        {
            'category': 'achievements',
            'title': 'Quantify Your Achievements',
            'description': 'Use numbers and metrics to show impact of your work',
            'action': 'Add specific metrics to at least 3 bullet points (e.g., "Improved performance by 40%")',
            'impact': 'medium'
        },
        {
            'category': 'skills',
            'title': 'Add Specific Technologies',
            'description': 'List specific tools and technologies you have experience with',
            'action': 'Create a "Technical Skills" section with categories',
            'impact': 'medium'
        }
    ],
    'optional_tips': [
        {
            'category': 'content',
            'title': 'Add Certifications',
            'description': 'Industry certifications can strengthen your profile',
            'action': 'Include relevant certifications if you have any (AWS, PMP, etc.)',
            'impact': 'low'
        }
    ]
}


class TestExcelExporter:
    """Test cases for ExcelExporter class."""
    
    def test_exporter_initialization(self):
        """Test ExcelExporter initialization."""
        exporter = ExcelExporter()
        assert exporter is not None
        assert exporter.workbook is None
    
    def test_export_jobs_basic(self):
        """Test basic job export without tips."""
        exporter = ExcelExporter()
        output = exporter.export_jobs(SAMPLE_JOBS, include_tips_sheet=False)
        
        assert output is not None
        assert isinstance(output, BytesIO)
        
        # Load and verify workbook
        wb = load_workbook(output)
        assert 'Jobs' in wb.sheetnames
        
        ws = wb['Jobs']
        assert ws['A1'].value == 'Job Title'
        assert ws['B1'].value == 'Company'
        assert ws['F1'].value == 'Score (%)'
    
    def test_export_jobs_with_data(self):
        """Test that job data is correctly exported."""
        output = export_jobs_to_excel(SAMPLE_JOBS, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Check first job data
        assert ws['A2'].value == 'Senior Python Developer'
        assert ws['B2'].value == 'Tech Corp'
        assert ws['C2'].value == 'San Francisco, CA'
        assert ws['F2'].value == 85
        assert ws['G2'].value == 'GREEN'
    
    def test_color_coding_green(self):
        """Test green color coding for high scores."""
        output = export_jobs_to_excel(SAMPLE_JOBS, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Check first row (score 85) has green fill
        cell = ws['A2']
        fill_color = cell.fill.start_color.rgb
        # Green color
        assert fill_color == 'FFCCFFCC' or fill_color == '00CCFFCC'
    
    def test_color_coding_yellow(self):
        """Test yellow color coding for medium scores."""
        output = export_jobs_to_excel(SAMPLE_JOBS, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Check second row (score 45) has yellow fill
        cell = ws['A3']
        fill_color = cell.fill.start_color.rgb
        # Yellow color
        assert fill_color == 'FFFFFF99' or fill_color == '00FFFF99'
    
    def test_color_coding_red(self):
        """Test red color coding for low scores."""
        output = export_jobs_to_excel(SAMPLE_JOBS, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Check third row (score 30) has red fill
        cell = ws['A4']
        fill_color = cell.fill.start_color.rgb
        # Red color
        assert fill_color == 'FFFFCCCC' or fill_color == '00FFCCCC'
    
    def test_export_with_tips_sheet(self):
        """Test export with separate tips sheet."""
        output = export_jobs_to_excel(SAMPLE_JOBS, SAMPLE_TIPS, include_tips_sheet=True)
        wb = load_workbook(output)
        
        assert 'Jobs' in wb.sheetnames
        assert 'Resume Tips' in wb.sheetnames
    
    def test_tips_sheet_structure(self):
        """Test structure of tips sheet."""
        output = export_jobs_to_excel(SAMPLE_JOBS, SAMPLE_TIPS, include_tips_sheet=True)
        wb = load_workbook(output)
        ws = wb['Resume Tips']
        
        # Check title
        assert 'Resume Optimization Tips' in ws['A1'].value
        
        # Check summary section
        assert ws['A3'].value == 'Summary:'
        assert SAMPLE_TIPS['summary'] in ws['A4'].value
        
        # Check headers exist
        assert ws['A11'].value == 'Priority'
        assert ws['B11'].value == 'Category'
        assert ws['C11'].value == 'Title'
    
    def test_tips_sheet_critical_tips(self):
        """Test critical tips are included in tips sheet."""
        output = export_jobs_to_excel(SAMPLE_JOBS, SAMPLE_TIPS, include_tips_sheet=True)
        wb = load_workbook(output)
        ws = wb['Resume Tips']
        
        # Find critical tips rows
        critical_found = False
        for row in range(1, 30):
            cell_value = ws[f'A{row}'].value
            if cell_value and '🔴 CRITICAL' in str(cell_value):
                critical_found = True
                break
        
        assert critical_found, "Critical tips not found in tips sheet"
    
    def test_tips_sheet_important_tips(self):
        """Test important tips are included in tips sheet."""
        output = export_jobs_to_excel(SAMPLE_JOBS, SAMPLE_TIPS, include_tips_sheet=True)
        wb = load_workbook(output)
        ws = wb['Resume Tips']
        
        # Find important tips rows
        important_found = False
        for row in range(1, 30):
            cell_value = ws[f'A{row}'].value
            if cell_value and '🟡 IMPORTANT' in str(cell_value):
                important_found = True
                break
        
        assert important_found, "Important tips not found in tips sheet"
    
    def test_tips_as_comments(self):
        """Test that tips are added as comments on Jobs sheet."""
        output = export_jobs_to_excel(SAMPLE_JOBS, SAMPLE_TIPS, include_tips_sheet=True)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Check that header cells have comments
        assert ws['A1'].comment is not None or ws['F1'].comment is not None
    
    def test_header_formatting(self):
        """Test header row formatting."""
        output = export_jobs_to_excel(SAMPLE_JOBS, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Check header has colored background
        header_cell = ws['A1']
        assert header_cell.fill.start_color.rgb is not None
        
        # Check header is bold
        assert header_cell.font.bold is True
    
    def test_freeze_panes(self):
        """Test that header row is frozen."""
        output = export_jobs_to_excel(SAMPLE_JOBS, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Check freeze panes is set
        assert ws.freeze_panes == 'A2'
    
    def test_auto_filter(self):
        """Test that auto-filter is enabled."""
        output = export_jobs_to_excel(SAMPLE_JOBS, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Check auto-filter is set
        assert ws.auto_filter.ref is not None
    
    def test_column_widths(self):
        """Test that column widths are set appropriately."""
        output = export_jobs_to_excel(SAMPLE_JOBS, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Check some column widths
        assert ws.column_dimensions['A'].width == 30  # Job Title
        assert ws.column_dimensions['H'].width == 40  # Description
    
    def test_export_empty_jobs_raises_error(self):
        """Test that exporting empty jobs list raises error."""
        exporter = ExcelExporter()
        
        with pytest.raises(ValueError, match="Cannot export empty jobs list"):
            exporter.export_jobs([])
    
    def test_export_to_file(self):
        """Test exporting to file on disk."""
        filename = 'test_export.xlsx'
        
        try:
            export_jobs_to_file(SAMPLE_JOBS, filename, include_tips_sheet=False)
            
            assert os.path.exists(filename)
            
            # Verify file can be opened
            wb = load_workbook(filename)
            assert 'Jobs' in wb.sheetnames
        
        finally:
            # Cleanup
            if os.path.exists(filename):
                os.remove(filename)
    
    def test_multiple_jobs(self):
        """Test export with multiple jobs."""
        # Create 10 jobs
        jobs = []
        for i in range(10):
            job = SAMPLE_JOBS[0].copy()
            job['title'] = f'Job {i+1}'
            job['score'] = {'overall_score': 50 + i*5, 'highlight': 'yellow'}
            jobs.append(job)
        
        output = export_jobs_to_excel(jobs, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Check that all jobs are exported (header + 10 jobs)
        assert ws.max_row == 11
    
    def test_jobs_without_scores(self):
        """Test export of jobs without score data."""
        jobs = [
            {
                'title': 'Test Job',
                'company': 'Test Company',
                'location': 'Test Location',
                'salary': '$100k',
                'job_type': 'Remote',
                'description': 'Test description',
                'link': 'http://test.com'
                # No score field
            }
        ]
        
        output = export_jobs_to_excel(jobs, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Should handle missing scores gracefully
        assert ws['A2'].value == 'Test Job'
        assert ws['F2'].value == 0  # Default score
    
    def test_long_description_truncation(self):
        """Test that long descriptions are truncated."""
        jobs = [SAMPLE_JOBS[0].copy()]
        jobs[0]['description'] = 'A' * 1000  # Very long description
        
        output = export_jobs_to_excel(jobs, include_tips_sheet=False)
        wb = load_workbook(output)
        ws = wb['Jobs']
        
        # Description should be truncated to 500 chars
        assert len(ws['H2'].value) <= 500
    
    def test_tips_without_optional_tips(self):
        """Test tips export when optional tips are empty."""
        tips = SAMPLE_TIPS.copy()
        tips['optional_tips'] = []
        
        output = export_jobs_to_excel(SAMPLE_JOBS, tips, include_tips_sheet=True)
        wb = load_workbook(output)
        
        # Should still work without errors
        assert 'Resume Tips' in wb.sheetnames
    
    def test_convenience_function(self):
        """Test convenience function export_jobs_to_excel."""
        output = export_jobs_to_excel(SAMPLE_JOBS)
        
        assert output is not None
        assert isinstance(output, BytesIO)
        
        wb = load_workbook(output)
        assert 'Jobs' in wb.sheetnames


def test_integration_with_real_data():
    """Integration test with realistic job data."""
    jobs = [
        {
            'title': 'Software Engineer',
            'company': 'Google',
            'location': 'Mountain View, CA',
            'salary': '$150,000 - $200,000',
            'job_type': 'Onsite',
            'description': 'We are looking for talented engineers...',
            'link': 'https://careers.google.com/jobs/1',
            'score': {
                'overall_score': 88,
                'highlight': 'green',
                'component_scores': {
                    'keyword_match': 92,
                    'salary_match': 85,
                    'location_match': 80,
                    'job_type_match': 90
                }
            }
        }
    ]
    
    tips = {
        'summary': 'Strong technical profile',
        'overall_assessment': {
            'strength_score': 85,
            'completeness': 'Excellent',
            'ats_compatibility': 'Good'
        },
        'critical_tips': [],
        'important_tips': [
            {
                'category': 'skills',
                'title': 'Add Cloud Skills',
                'description': 'Add AWS or GCP experience',
                'action': 'Include cloud projects',
                'impact': 'medium'
            }
        ],
        'optional_tips': []
    }
    
    output = export_jobs_to_excel(jobs, tips, include_tips_sheet=True)
    wb = load_workbook(output)
    
    assert 'Jobs' in wb.sheetnames
    assert 'Resume Tips' in wb.sheetnames
    
    # Verify data
    ws = wb['Jobs']
    assert ws['A2'].value == 'Software Engineer'
    assert ws['B2'].value == 'Google'


if __name__ == '__main__':
    # Run tests
    pytest.main([__file__, '-v'])

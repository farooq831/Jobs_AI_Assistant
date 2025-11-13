"""
Excel Upload Module for Job Application Status Tracking

This module handles:
- Parsing uploaded Excel sheets for job application status changes
- Validating data integrity on import
- Supporting status updates: Applied, Interview, Offer, Rejected, Pending
- Tracking status history and timestamps
- Error detection and reporting

Author: AI Job Application Assistant
Date: November 2025
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import os
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelUploadError(Exception):
    """Custom exception for Excel upload errors"""
    pass


class ExcelUploader:
    """
    Handles parsing and validation of uploaded Excel files for job application status tracking
    """
    
    # Valid application statuses
    VALID_STATUSES = {
        'Applied', 'Interview', 'Offer', 'Rejected', 'Pending',
        'applied', 'interview', 'offer', 'rejected', 'pending',
        'APPLIED', 'INTERVIEW', 'OFFER', 'REJECTED', 'PENDING'
    }
    
    # Normalized status mapping
    STATUS_NORMALIZATION = {
        'applied': 'Applied',
        'interview': 'Interview',
        'offer': 'Offer',
        'rejected': 'Rejected',
        'pending': 'Pending'
    }
    
    # Required fields for job data
    REQUIRED_FIELDS = ['job_id', 'title', 'company']
    
    # Expected column names (case-insensitive)
    EXPECTED_COLUMNS = {
        'job_id': ['job_id', 'id', 'job id', 'jobid'],
        'title': ['title', 'job_title', 'job title', 'position'],
        'company': ['company', 'company_name', 'company name', 'employer'],
        'status': ['status', 'application_status', 'application status', 'app_status'],
        'applied_date': ['applied_date', 'applied date', 'date_applied', 'date applied', 'application_date'],
        'notes': ['notes', 'comments', 'note', 'comment', 'remarks']
    }
    
    def __init__(self):
        """Initialize the Excel uploader"""
        self.validation_errors = []
        self.validation_warnings = []
        self.parsed_data = []
        
    def parse_excel_file(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Parse an Excel file and extract job application status data
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to parse (defaults to first sheet)
            
        Returns:
            Dict with parsed data and validation results
            
        Raises:
            ExcelUploadError: If file cannot be parsed or is invalid
        """
        self.validation_errors = []
        self.validation_warnings = []
        self.parsed_data = []
        
        try:
            # Load workbook
            if not os.path.exists(file_path):
                raise ExcelUploadError(f"File not found: {file_path}")
                
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Select sheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ExcelUploadError(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
                
            logger.info(f"Parsing sheet: {sheet.title}")
            
            # Parse the sheet
            result = self._parse_sheet(sheet)
            
            workbook.close()
            
            return result
            
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise ExcelUploadError(f"Invalid Excel file: {str(e)}")
        except Exception as e:
            raise ExcelUploadError(f"Error parsing Excel file: {str(e)}")
            
    def parse_excel_bytes(self, file_bytes: bytes, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Parse Excel file from bytes (for uploaded files)
        
        Args:
            file_bytes: Excel file content as bytes
            sheet_name: Name of the sheet to parse
            
        Returns:
            Dict with parsed data and validation results
        """
        import io
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            
            # Select sheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ExcelUploadError(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
                
            logger.info(f"Parsing sheet: {sheet.title}")
            
            # Parse the sheet
            result = self._parse_sheet(sheet)
            
            workbook.close()
            
            return result
            
        except Exception as e:
            raise ExcelUploadError(f"Error parsing Excel bytes: {str(e)}")
            
    def _parse_sheet(self, sheet) -> Dict[str, Any]:
        """
        Internal method to parse a worksheet
        
        Args:
            sheet: openpyxl worksheet object
            
        Returns:
            Dict with parsed data and validation results
        """
        # Find header row
        header_row = self._find_header_row(sheet)
        if header_row is None:
            raise ExcelUploadError("Could not find header row in Excel sheet")
            
        # Map columns
        column_mapping = self._map_columns(sheet, header_row)
        if not column_mapping:
            raise ExcelUploadError("Could not identify required columns in Excel sheet")
            
        # Validate required columns
        missing_required = []
        for field in self.REQUIRED_FIELDS:
            if field not in column_mapping:
                missing_required.append(field)
                
        if missing_required:
            raise ExcelUploadError(f"Missing required columns: {', '.join(missing_required)}")
            
        # Parse data rows
        data_rows = []
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row_data = self._parse_row(sheet, row_idx, column_mapping)
            if row_data:
                data_rows.append(row_data)
                
        self.parsed_data = data_rows
        
        # Generate summary
        summary = self._generate_summary(data_rows)
        
        return {
            'success': len(self.validation_errors) == 0,
            'data': data_rows,
            'total_rows': len(data_rows),
            'valid_rows': len([r for r in data_rows if r.get('valid', True)]),
            'errors': self.validation_errors,
            'warnings': self.validation_warnings,
            'summary': summary,
            'column_mapping': column_mapping
        }
        
    def _find_header_row(self, sheet) -> Optional[int]:
        """
        Find the header row in the worksheet
        
        Returns:
            Row index (1-based) or None if not found
        """
        # Check first 10 rows for headers
        for row_idx in range(1, min(11, sheet.max_row + 1)):
            row_values = [cell.value for cell in sheet[row_idx] if cell.value]
            row_values_lower = [str(v).lower().strip() for v in row_values]
            
            # Check if this row contains expected column names
            matches = 0
            for field, variations in self.EXPECTED_COLUMNS.items():
                if any(var in row_values_lower for var in variations):
                    matches += 1
                    
            # If we find at least 2 expected columns, it's likely the header
            if matches >= 2:
                return row_idx
                
        return None
        
    def _map_columns(self, sheet, header_row: int) -> Dict[str, int]:
        """
        Map column names to column indices
        
        Args:
            sheet: Worksheet object
            header_row: Header row index
            
        Returns:
            Dict mapping field names to column indices (0-based)
        """
        mapping = {}
        header_cells = list(sheet[header_row])
        
        for col_idx, cell in enumerate(header_cells):
            if not cell.value:
                continue
                
            col_name = str(cell.value).lower().strip()
            
            # Try to match with expected columns
            for field, variations in self.EXPECTED_COLUMNS.items():
                if col_name in variations:
                    mapping[field] = col_idx
                    break
                    
        return mapping
        
    def _parse_row(self, sheet, row_idx: int, column_mapping: Dict[str, int]) -> Optional[Dict[str, Any]]:
        """
        Parse a single data row
        
        Args:
            sheet: Worksheet object
            row_idx: Row index (1-based)
            column_mapping: Column mapping dict
            
        Returns:
            Dict with row data or None if row is empty
        """
        row = list(sheet[row_idx])
        
        # Check if row is empty
        if all(cell.value is None for cell in row):
            return None
            
        row_data = {
            'row_number': row_idx,
            'valid': True,
            'errors': []
        }
        
        # Extract mapped fields
        for field, col_idx in column_mapping.items():
            if col_idx < len(row):
                value = row[col_idx].value
                
                # Process the value based on field type
                if field == 'status' and value:
                    value = self._normalize_status(value, row_idx)
                elif field == 'applied_date' and value:
                    value = self._parse_date(value, row_idx)
                elif value is not None:
                    value = str(value).strip()
                    
                row_data[field] = value
            else:
                row_data[field] = None
                
        # Validate row
        self._validate_row(row_data)
        
        return row_data
        
    def _normalize_status(self, status: Any, row_idx: int) -> Optional[str]:
        """
        Normalize status value
        
        Args:
            status: Raw status value
            row_idx: Row index for error reporting
            
        Returns:
            Normalized status or None
        """
        if status is None:
            return None
            
        status_str = str(status).strip()
        status_lower = status_str.lower()
        
        # Check if valid
        if status_str not in self.VALID_STATUSES:
            self.validation_warnings.append({
                'row': row_idx,
                'field': 'status',
                'message': f"Invalid status '{status_str}'. Valid values: {', '.join(sorted(set(self.STATUS_NORMALIZATION.values())))}"
            })
            return None
            
        # Normalize
        return self.STATUS_NORMALIZATION.get(status_lower, status_str)
        
    def _parse_date(self, date_value: Any, row_idx: int) -> Optional[str]:
        """
        Parse and validate date value
        
        Args:
            date_value: Raw date value
            row_idx: Row index for error reporting
            
        Returns:
            ISO format date string or None
        """
        if date_value is None:
            return None
            
        try:
            # If already a datetime object
            if isinstance(date_value, datetime):
                return date_value.strftime('%Y-%m-%d')
                
            # Try to parse string
            date_str = str(date_value).strip()
            
            # Try common date formats
            formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d', '%m-%d-%Y', '%d-%m-%Y']
            for fmt in formats:
                try:
                    parsed_date = datetime.strptime(date_str, fmt)
                    return parsed_date.strftime('%Y-%m-%d')
                except ValueError:
                    continue
                    
            # If no format worked
            self.validation_warnings.append({
                'row': row_idx,
                'field': 'applied_date',
                'message': f"Could not parse date '{date_str}'. Expected format: YYYY-MM-DD"
            })
            return None
            
        except Exception as e:
            self.validation_warnings.append({
                'row': row_idx,
                'field': 'applied_date',
                'message': f"Error parsing date: {str(e)}"
            })
            return None
            
    def _validate_row(self, row_data: Dict[str, Any]) -> None:
        """
        Validate a parsed row
        
        Args:
            row_data: Parsed row data (modified in place)
        """
        errors = []
        
        # Check required fields
        for field in self.REQUIRED_FIELDS:
            if not row_data.get(field):
                errors.append(f"Missing required field: {field}")
                
        # Additional validations
        if row_data.get('job_id'):
            job_id = str(row_data['job_id']).strip()
            if len(job_id) == 0:
                errors.append("job_id cannot be empty")
                
        if errors:
            row_data['valid'] = False
            row_data['errors'] = errors
            self.validation_errors.extend([{
                'row': row_data['row_number'],
                'errors': errors
            }])
            
    def _generate_summary(self, data_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Generate summary statistics for parsed data
        
        Args:
            data_rows: List of parsed rows
            
        Returns:
            Summary statistics dict
        """
        summary = {
            'total_jobs': len(data_rows),
            'valid_jobs': len([r for r in data_rows if r.get('valid', True)]),
            'invalid_jobs': len([r for r in data_rows if not r.get('valid', True)]),
            'status_counts': {},
            'jobs_with_dates': 0,
            'jobs_with_notes': 0
        }
        
        # Count statuses
        for row in data_rows:
            if row.get('valid', True):
                status = row.get('status')
                if status:
                    summary['status_counts'][status] = summary['status_counts'].get(status, 0) + 1
                    
                if row.get('applied_date'):
                    summary['jobs_with_dates'] += 1
                    
                if row.get('notes'):
                    summary['jobs_with_notes'] += 1
                    
        return summary
        
    def validate_against_stored_jobs(self, parsed_data: List[Dict[str, Any]], 
                                     stored_jobs: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Validate parsed data against stored jobs to check for matches and discrepancies
        
        Args:
            parsed_data: List of parsed job records from Excel
            stored_jobs: List of jobs from storage
            
        Returns:
            Validation results with matches, new jobs, and mismatches
        """
        results = {
            'matched_jobs': [],
            'new_jobs': [],
            'mismatched_jobs': [],
            'total_parsed': len(parsed_data),
            'total_stored': len(stored_jobs)
        }
        
        # Create lookup for stored jobs
        stored_lookup = {str(job.get('job_id')): job for job in stored_jobs}
        
        for parsed_job in parsed_data:
            if not parsed_job.get('valid', True):
                continue
                
            job_id = str(parsed_job.get('job_id'))
            
            if job_id in stored_lookup:
                # Found matching job
                stored_job = stored_lookup[job_id]
                match_info = {
                    'job_id': job_id,
                    'parsed': parsed_job,
                    'stored': stored_job,
                    'discrepancies': []
                }
                
                # Check for discrepancies
                if parsed_job.get('title') and parsed_job['title'] != stored_job.get('title'):
                    match_info['discrepancies'].append({
                        'field': 'title',
                        'parsed': parsed_job['title'],
                        'stored': stored_job.get('title')
                    })
                    
                if parsed_job.get('company') and parsed_job['company'] != stored_job.get('company'):
                    match_info['discrepancies'].append({
                        'field': 'company',
                        'parsed': parsed_job['company'],
                        'stored': stored_job.get('company')
                    })
                    
                if match_info['discrepancies']:
                    results['mismatched_jobs'].append(match_info)
                else:
                    results['matched_jobs'].append(match_info)
            else:
                # New job not in storage
                results['new_jobs'].append(parsed_job)
                
        return results
        
    def get_status_updates(self, parsed_data: List[Dict[str, Any]], 
                          stored_jobs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Extract status updates from parsed data for existing jobs
        
        Args:
            parsed_data: List of parsed job records
            stored_jobs: List of jobs from storage
            
        Returns:
            List of status update records
        """
        updates = []
        stored_lookup = {str(job.get('job_id')): job for job in stored_jobs}
        
        for parsed_job in parsed_data:
            if not parsed_job.get('valid', True):
                continue
                
            job_id = str(parsed_job.get('job_id'))
            new_status = parsed_job.get('status')
            
            if not new_status:
                continue
                
            if job_id in stored_lookup:
                stored_job = stored_lookup[job_id]
                old_status = stored_job.get('application_status')
                
                # Only include if status changed
                if new_status != old_status:
                    updates.append({
                        'job_id': job_id,
                        'old_status': old_status,
                        'new_status': new_status,
                        'applied_date': parsed_job.get('applied_date'),
                        'notes': parsed_job.get('notes'),
                        'timestamp': datetime.now().isoformat()
                    })
                    
        return updates


if __name__ == "__main__":
    # Example usage
    print("Excel Uploader Module")
    print("=" * 50)
    print("\nThis module handles parsing and validation of uploaded Excel files")
    print("for job application status tracking.")
    print("\nValid statuses:", ", ".join(sorted(set(ExcelUploader.STATUS_NORMALIZATION.values()))))
    print("\nRequired fields:", ", ".join(ExcelUploader.REQUIRED_FIELDS))

"""
Excel Export Module
Export job listings with scores and color-coded highlights using openpyxl.
Includes resume optimization tips as comments or separate sheet.
"""

import logging
from typing import Dict, List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from io import BytesIO

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Export job listings to Excel with formatting and color-coding based on scores.
    """
    
    # Color definitions for highlighting
    COLORS = {
        'red': 'FFCCCC',      # Poor match (< 40%)
        'yellow': 'FFFF99',   # Fair match (40-70%)
        'white': 'FFFFFF',    # Good match (> 70%)
        'green': 'CCFFCC',    # Excellent match (> 85%)
        'header': '4472C4',   # Header background
        'tips_header': '70AD47'  # Tips sheet header
    }
    
    # Column widths for better readability
    COLUMN_WIDTHS = {
        'A': 30,  # Job Title
        'B': 25,  # Company
        'C': 20,  # Location
        'D': 15,  # Salary
        'E': 12,  # Job Type
        'F': 12,  # Score
        'G': 12,  # Highlight
        'H': 40,  # Description
        'I': 30   # Link
    }
    
    def __init__(self):
        """Initialize the Excel exporter."""
        self.workbook = None
        logger.info("ExcelExporter initialized")
    
    def export_jobs(self, jobs: List[Dict], 
                    resume_tips: Optional[Dict] = None,
                    filename: Optional[str] = None,
                    include_tips_sheet: bool = True) -> BytesIO:
        """
        Export jobs list to Excel with formatting.
        
        Args:
            jobs: List of job dictionaries with scores and highlights
            resume_tips: Optional resume optimization tips dictionary
            filename: Optional custom filename
            include_tips_sheet: Whether to include a separate tips sheet
            
        Returns:
            BytesIO object containing the Excel file
        """
        if not jobs:
            logger.warning("No jobs to export")
            raise ValueError("Cannot export empty jobs list")
        
        # Create workbook
        self.workbook = Workbook()
        
        # Create jobs sheet
        self._create_jobs_sheet(jobs, resume_tips)
        
        # Create tips sheet if requested and tips available
        if include_tips_sheet and resume_tips:
            self._create_tips_sheet(resume_tips)
        
        # Remove default sheet if we created named sheets
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        logger.info(f"Exported {len(jobs)} jobs to Excel")
        return output
    
    def _create_jobs_sheet(self, jobs: List[Dict], resume_tips: Optional[Dict] = None):
        """
        Create the main jobs listing sheet with color-coding.
        
        Args:
            jobs: List of job dictionaries
            resume_tips: Optional tips to add as cell comments
        """
        # Get or create Jobs sheet
        if 'Jobs' in self.workbook.sheetnames:
            ws = self.workbook['Jobs']
        else:
            ws = self.workbook.active
            ws.title = 'Jobs'
        
        # Define headers
        headers = [
            'Job Title', 'Company', 'Location', 'Salary', 
            'Job Type', 'Score (%)', 'Match Quality', 
            'Description', 'Link'
        ]
        
        # Write headers with formatting
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color=self.COLORS['header'], 
                                   end_color=self.COLORS['header'], 
                                   fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        
        # Write job data with formatting
        for row_num, job in enumerate(jobs, 2):
            self._write_job_row(ws, row_num, job)
        
        # Add resume tips as comments on header row if available
        if resume_tips:
            self._add_tips_as_comments(ws, resume_tips)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Auto-filter
        ws.auto_filter.ref = f'A1:I{len(jobs) + 1}'
    
    def _write_job_row(self, worksheet, row_num: int, job: Dict):
        """
        Write a single job row with appropriate formatting.
        
        Args:
            worksheet: Excel worksheet object
            row_num: Row number to write to
            job: Job dictionary
        """
        # Extract job data
        title = job.get('title', 'N/A')
        company = job.get('company', 'N/A')
        location = job.get('location', 'N/A')
        salary = job.get('salary', 'N/A')
        job_type = job.get('job_type', 'N/A')
        
        # Get score data
        score_data = job.get('score', {})
        if isinstance(score_data, dict):
            overall_score = score_data.get('overall_score', 0)
            highlight = score_data.get('highlight', 'white')
        else:
            overall_score = 0
            highlight = 'white'
        
        description = job.get('description', 'N/A')
        link = job.get('link') or job.get('url', 'N/A')  # Support both 'link' and 'url' fields
        
        # Determine row fill color based on highlight
        fill_color = self._get_fill_color(highlight, overall_score)
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        # Create border style
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write cells
        cells_data = [
            (1, title),
            (2, company),
            (3, location),
            (4, salary),
            (5, job_type),
            (6, overall_score),
            (7, highlight.upper()),
            (8, description[:500] if len(description) > 500 else description),  # Truncate long descriptions
            (9, link)
        ]
        
        for col_num, value in cells_data:
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Make score bold
            if col_num == 6:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Make highlight bold and centered
            if col_num == 7:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Make link clickable and blue
            if col_num == 9 and value != 'N/A' and value.startswith('http'):
                cell.hyperlink = value
                cell.font = Font(color='0563C1', underline='single')
                cell.value = 'View Job'  # Display text instead of full URL
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _get_fill_color(self, highlight: str, score: float) -> str:
        """
        Determine fill color based on highlight and score.
        
        Args:
            highlight: Highlight category (red, yellow, white)
            score: Overall match score
            
        Returns:
            Hex color code
        """
        highlight = highlight.lower()
        
        # Use green for excellent matches (>85%)
        if score >= 85:
            return self.COLORS['green']
        
        # Use defined colors for highlights
        if highlight in self.COLORS:
            return self.COLORS[highlight]
        
        # Default to white
        return self.COLORS['white']
    
    def _add_tips_as_comments(self, worksheet, tips: Dict):
        """
        Add resume optimization tips as comments on the header row.
        
        Args:
            worksheet: Excel worksheet object
            tips: Resume tips dictionary
        """
        # Add summary as comment on Job Title header
        if 'summary' in tips:
            cell = worksheet['A1']
            comment_text = f"Resume Summary:\n{tips['summary'][:200]}"
            cell.comment = Comment(comment_text, 'Resume Analyzer')
        
        # Add critical tips count as comment on Score header
        critical_count = len(tips.get('critical_tips', []))
        important_count = len(tips.get('important_tips', []))
        
        cell = worksheet['F1']
        comment_text = (
            f"Resume Optimization:\n"
            f"🔴 {critical_count} Critical tips\n"
            f"🟡 {important_count} Important tips\n"
            f"See 'Resume Tips' sheet for details"
        )
        cell.comment = Comment(comment_text, 'Resume Analyzer')
    
    def _create_tips_sheet(self, tips: Dict):
        """
        Create a separate sheet for resume optimization tips.
        
        Args:
            tips: Resume tips dictionary
        """
        ws = self.workbook.create_sheet(title='Resume Tips')
        
        # Add title
        ws['A1'] = 'Resume Optimization Tips'
        title_font = Font(bold=True, size=16, color='FFFFFF')
        title_fill = PatternFill(start_color=self.COLORS['tips_header'],
                                 end_color=self.COLORS['tips_header'],
                                 fill_type='solid')
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws.merge_cells('A1:F1')
        
        # Add summary
        ws['A3'] = 'Summary:'
        ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = tips.get('summary', 'No summary available')
        ws.merge_cells('A4:F4')
        ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Add overall assessment
        row = 6
        ws[f'A{row}'] = 'Overall Assessment:'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        assessment = tips.get('overall_assessment', {})
        row += 1
        ws[f'A{row}'] = f"Strength Score: {assessment.get('strength_score', 'N/A')}/100"
        row += 1
        ws[f'A{row}'] = f"Completeness: {assessment.get('completeness', 'N/A')}"
        row += 1
        ws[f'A{row}'] = f"ATS Compatibility: {assessment.get('ats_compatibility', 'N/A')}"
        row += 1
        
        # Add headers for tips table
        row += 1
        headers = ['Priority', 'Category', 'Title', 'Description', 'Action', 'Impact']
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color=self.COLORS['tips_header'],
                                  end_color=self.COLORS['tips_header'],
                                  fill_type='solid')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 12
        
        # Add tips data
        row += 1
        
        # Add critical tips
        for tip in tips.get('critical_tips', []):
            self._write_tip_row(ws, row, '🔴 CRITICAL', tip, 'FFCCCC')
            row += 1
        
        # Add important tips
        for tip in tips.get('important_tips', []):
            self._write_tip_row(ws, row, '🟡 IMPORTANT', tip, 'FFFF99')
            row += 1
        
        # Add optional tips
        for tip in tips.get('optional_tips', []):
            self._write_tip_row(ws, row, '⚪ OPTIONAL', tip, 'FFFFFF')
            row += 1
        
        # Freeze header rows
        ws.freeze_panes = f'A{row - len(tips.get("critical_tips", [])) - len(tips.get("important_tips", [])) - len(tips.get("optional_tips", []))}'
    
    def _write_tip_row(self, worksheet, row_num: int, priority: str, tip: Dict, fill_color: str):
        """
        Write a single tip row with formatting.
        
        Args:
            worksheet: Excel worksheet object
            row_num: Row number to write to
            priority: Priority level string
            tip: Tip dictionary
            fill_color: Background color for the row
        """
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        cells_data = [
            (1, priority),
            (2, tip.get('category', 'N/A').upper()),
            (3, tip.get('title', 'N/A')),
            (4, tip.get('description', 'N/A')),
            (5, tip.get('action', 'N/A')),
            (6, tip.get('impact', 'N/A').upper())
        ]
        
        for col_num, value in cells_data:
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Make priority bold
            if col_num == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def export_jobs_to_file(self, jobs: List[Dict], 
                           filename: str,
                           resume_tips: Optional[Dict] = None,
                           include_tips_sheet: bool = True):
        """
        Export jobs to an Excel file on disk.
        
        Args:
            jobs: List of job dictionaries
            filename: Output filename
            resume_tips: Optional resume tips
            include_tips_sheet: Whether to include tips sheet
        """
        output = self.export_jobs(jobs, resume_tips, filename, include_tips_sheet)
        
        with open(filename, 'wb') as f:
            f.write(output.getvalue())
        
        logger.info(f"Exported jobs to file: {filename}")


def export_jobs_to_excel(jobs: List[Dict], 
                         resume_tips: Optional[Dict] = None,
                         filename: Optional[str] = None,
                         include_tips_sheet: bool = True) -> BytesIO:
    """
    Convenience function to export jobs to Excel.
    
    Args:
        jobs: List of job dictionaries with scores and highlights
        resume_tips: Optional resume optimization tips
        filename: Optional custom filename (not used for BytesIO output)
        include_tips_sheet: Whether to include separate tips sheet
        
    Returns:
        BytesIO object containing the Excel file
    """
    exporter = ExcelExporter()
    return exporter.export_jobs(jobs, resume_tips, filename, include_tips_sheet)


def export_jobs_to_file(jobs: List[Dict],
                       filename: str,
                       resume_tips: Optional[Dict] = None,
                       include_tips_sheet: bool = True):
    """
    Convenience function to export jobs to Excel file on disk.
    
    Args:
        jobs: List of job dictionaries with scores and highlights
        filename: Output filename
        resume_tips: Optional resume optimization tips
        include_tips_sheet: Whether to include separate tips sheet
    """
    exporter = ExcelExporter()
    exporter.export_jobs_to_file(jobs, filename, resume_tips, include_tips_sheet)


if __name__ == '__main__':
    # Example usage
    sample_jobs = [
        {
            'title': 'Senior Python Developer',
            'company': 'Tech Corp',
            'location': 'San Francisco, CA',
            'salary': '$120k-$150k',
            'job_type': 'Remote',
            'description': 'Looking for an experienced Python developer...',
            'link': 'https://example.com/job1',
            'score': {
                'overall_score': 85,
                'highlight': 'green'
            }
        },
        {
            'title': 'Junior Developer',
            'company': 'Startup Inc',
            'location': 'New York, NY',
            'salary': '$60k-$80k',
            'job_type': 'Onsite',
            'description': 'Entry-level position for new graduates...',
            'link': 'https://example.com/job2',
            'score': {
                'overall_score': 45,
                'highlight': 'yellow'
            }
        }
    ]
    
    sample_tips = {
        'summary': 'Your resume shows good technical skills but needs improvement in quantifying achievements.',
        'overall_assessment': {
            'strength_score': 72,
            'completeness': 'Good',
            'ats_compatibility': 'Fair'
        },
        'critical_tips': [
            {
                'category': 'keywords',
                'title': 'Add Missing Technical Keywords',
                'description': 'Your resume is missing important keywords like Python, AWS, Docker',
                'action': 'Add these keywords naturally in your experience section',
                'impact': 'high'
            }
        ],
        'important_tips': [
            {
                'category': 'achievements',
                'title': 'Quantify Your Achievements',
                'description': 'Use numbers and metrics to show impact',
                'action': 'Add specific metrics to at least 3 bullet points',
                'impact': 'medium'
            }
        ],
        'optional_tips': []
    }
    
    print("Exporting sample jobs to Excel...")
    export_jobs_to_file(sample_jobs, 'sample_jobs_export.xlsx', sample_tips, include_tips_sheet=True)
    print("Export complete! Check sample_jobs_export.xlsx")

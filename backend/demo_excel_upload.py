"""
Demo Script for Excel Upload Module (Task 7.3)

This script demonstrates the Excel upload functionality for job application status tracking:
- Creating sample Excel files
- Parsing Excel files
- Validating data
- Applying status updates
- Integration with storage manager

Run with: python3 demo_excel_upload.py
"""

import os
import sys
from datetime import datetime, timedelta

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Note: openpyxl not installed. Install with: pip install openpyxl")

from excel_uploader import ExcelUploader, ExcelUploadError
from storage_manager import JobStorageManager


def print_section(title):
    """Print a section header"""
    print("\n" + "=" * 70)
    print(f" {title}")
    print("=" * 70)


def create_sample_excel(filename):
    """Create a sample Excel file for demonstration"""
    if not OPENPYXL_AVAILABLE:
        print("Cannot create sample Excel file - openpyxl not installed")
        return False
    
    print(f"\nCreating sample Excel file: {filename}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Applications"
    
    # Headers
    headers = ['job_id', 'title', 'company', 'status', 'applied_date', 'notes']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Sample data
    today = datetime.now()
    data = [
        ['job1', 'Software Engineer', 'Tech Corp', 'Applied', 
         (today - timedelta(days=5)).strftime('%Y-%m-%d'), 'Applied via LinkedIn'],
        ['job2', 'Data Analyst', 'Data Inc', 'Interview', 
         (today - timedelta(days=3)).strftime('%Y-%m-%d'), 'Phone screening completed'],
        ['job3', 'Product Manager', 'Product Co', 'Pending', 
         '', 'Preparing application'],
        ['job4', 'DevOps Engineer', 'Cloud Systems', 'Applied', 
         (today - timedelta(days=7)).strftime('%Y-%m-%d'), 'Applied via company website'],
        ['job5', 'UX Designer', 'Design Studio', 'Offer', 
         (today - timedelta(days=10)).strftime('%Y-%m-%d'), 'Received offer letter!'],
    ]
    
    for row in data:
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    wb.close()
    
    print(f"✓ Created sample file with {len(data)} job applications")
    return True


def demo_parse_excel(filename):
    """Demonstrate parsing an Excel file"""
    print_section("Demo 1: Parse Excel File")
    
    if not os.path.exists(filename):
        print(f"Error: File not found: {filename}")
        return None
    
    uploader = ExcelUploader()
    
    try:
        result = uploader.parse_excel_file(filename)
        
        print(f"\n✓ Parsing successful!")
        print(f"  Total rows: {result['total_rows']}")
        print(f"  Valid rows: {result['valid_rows']}")
        print(f"  Errors: {len(result['errors'])}")
        print(f"  Warnings: {len(result['warnings'])}")
        
        print(f"\nSummary:")
        summary = result['summary']
        print(f"  Total jobs: {summary['total_jobs']}")
        print(f"  Status breakdown:")
        for status, count in summary['status_counts'].items():
            print(f"    - {status}: {count}")
        print(f"  Jobs with dates: {summary['jobs_with_dates']}")
        print(f"  Jobs with notes: {summary['jobs_with_notes']}")
        
        if result['warnings']:
            print(f"\nWarnings:")
            for warning in result['warnings'][:3]:  # Show first 3
                print(f"  - Row {warning['row']}: {warning['message']}")
        
        return result
        
    except ExcelUploadError as e:
        print(f"✗ Parsing failed: {e}")
        return None


def demo_validate_against_storage(parse_result):
    """Demonstrate validation against stored jobs"""
    print_section("Demo 2: Validate Against Stored Jobs")
    
    if not parse_result or not parse_result['success']:
        print("Skipping - no valid parse result")
        return
    
    # Create temporary storage with some jobs
    storage_manager = JobStorageManager(storage_dir='demo_data')
    
    # Add some jobs to storage
    print("\nAdding sample jobs to storage...")
    sample_jobs = [
        {
            'job_id': 'job1',
            'title': 'Software Engineer',
            'company': 'Tech Corp',
            'location': 'San Francisco',
            'salary_min': 100000,
            'salary_max': 150000
        },
        {
            'job_id': 'job2',
            'title': 'Data Analyst',
            'company': 'Data Inc',
            'location': 'New York',
            'salary_min': 80000,
            'salary_max': 120000
        },
        {
            'job_id': 'job3',
            'title': 'Product Manager',
            'company': 'Product Co',
            'location': 'Remote',
            'salary_min': 120000,
            'salary_max': 160000
        }
    ]
    
    for job in sample_jobs:
        storage_manager.add_job(job)
    
    print(f"✓ Added {len(sample_jobs)} jobs to storage")
    
    # Validate
    uploader = ExcelUploader()
    stored_jobs = storage_manager.get_all_jobs()
    validation_result = uploader.validate_against_stored_jobs(
        parse_result['data'],
        stored_jobs
    )
    
    print(f"\nValidation Results:")
    print(f"  Total parsed: {validation_result['total_parsed']}")
    print(f"  Total stored: {validation_result['total_stored']}")
    print(f"  Matched jobs: {len(validation_result['matched_jobs'])}")
    print(f"  New jobs: {len(validation_result['new_jobs'])}")
    print(f"  Mismatched jobs: {len(validation_result['mismatched_jobs'])}")
    
    if validation_result['matched_jobs']:
        print(f"\nMatched jobs:")
        for match in validation_result['matched_jobs'][:3]:
            print(f"  - {match['job_id']}: {match['parsed']['title']}")
    
    if validation_result['new_jobs']:
        print(f"\nNew jobs (not in storage):")
        for job in validation_result['new_jobs']:
            print(f"  - {job['job_id']}: {job['title']} at {job['company']}")
    
    if validation_result['mismatched_jobs']:
        print(f"\nMismatched jobs:")
        for mismatch in validation_result['mismatched_jobs']:
            print(f"  - {mismatch['job_id']}:")
            for disc in mismatch['discrepancies']:
                print(f"    {disc['field']}: '{disc['parsed']}' vs '{disc['stored']}'")


def demo_status_updates(parse_result):
    """Demonstrate status update extraction and application"""
    print_section("Demo 3: Extract and Apply Status Updates")
    
    if not parse_result or not parse_result['success']:
        print("Skipping - no valid parse result")
        return
    
    # Create storage manager
    storage_manager = JobStorageManager(storage_dir='demo_data')
    
    # Get status updates
    uploader = ExcelUploader()
    stored_jobs = storage_manager.get_all_jobs()
    status_updates = uploader.get_status_updates(parse_result['data'], stored_jobs)
    
    print(f"\nFound {len(status_updates)} status updates")
    
    if status_updates:
        print(f"\nStatus updates to apply:")
        for update in status_updates:
            old_status = update.get('old_status', 'None')
            new_status = update.get('new_status')
            print(f"  - Job {update['job_id']}: {old_status} → {new_status}")
            if update.get('notes'):
                print(f"    Note: {update['notes']}")
        
        # Apply updates
        print(f"\nApplying updates...")
        result = storage_manager.batch_update_job_statuses(status_updates)
        
        if result['success']:
            print(f"✓ Successfully updated {result['updated']} job(s)")
            if result['not_found'] > 0:
                print(f"  {result['not_found']} job(s) not found in storage")
        else:
            print(f"✗ Update failed: {result.get('error', 'Unknown error')}")
    else:
        print("No status changes detected")


def demo_status_summary():
    """Demonstrate status summary"""
    print_section("Demo 4: Application Status Summary")
    
    storage_manager = JobStorageManager(storage_dir='demo_data')
    summary = storage_manager.get_status_summary()
    
    print(f"\nApplication Status Summary:")
    print(f"  Total jobs: {summary['total_jobs']}")
    print(f"  Jobs with status: {summary['jobs_with_status']}")
    print(f"  Jobs without status: {summary['jobs_without_status']}")
    
    if summary['status_counts']:
        print(f"\n  Status Breakdown:")
        for status, count in summary['status_counts'].items():
            percentage = (count / summary['total_jobs'] * 100) if summary['total_jobs'] > 0 else 0
            print(f"    - {status}: {count} ({percentage:.1f}%)")
    
    if summary.get('recent_applications'):
        print(f"\n  Recent Applications:")
        for app in summary['recent_applications'][:5]:
            date = app.get('applied_date', 'N/A')
            status = app.get('status', 'N/A')
            print(f"    - {app['title']} at {app['company']}")
            print(f"      Applied: {date}, Status: {status}")


def demo_single_status_update():
    """Demonstrate updating a single job status"""
    print_section("Demo 5: Update Single Job Status")
    
    storage_manager = JobStorageManager(storage_dir='demo_data')
    
    # Update a job status
    job_id = 'job4'
    print(f"\nUpdating status for job {job_id}...")
    
    result = storage_manager.update_job_status(
        job_id=job_id,
        status='Interview',
        applied_date=datetime.now().strftime('%Y-%m-%d'),
        notes='Received invitation for technical interview'
    )
    
    if result['success']:
        print(f"✓ Status updated successfully")
        
        # Get updated job
        jobs = storage_manager.get_all_jobs()
        job = next((j for j in jobs if j['job_id'] == job_id), None)
        
        if job:
            print(f"\nUpdated job details:")
            print(f"  Job ID: {job['job_id']}")
            print(f"  Title: {job['title']}")
            print(f"  Company: {job['company']}")
            print(f"  Status: {job.get('application_status', 'N/A')}")
            print(f"  Applied: {job.get('applied_date', 'N/A')}")
            print(f"  Notes: {job.get('application_notes', 'N/A')}")
            
            if 'status_history' in job:
                print(f"\n  Status History:")
                for i, entry in enumerate(job['status_history'], 1):
                    old = entry.get('old_status', 'None')
                    new = entry.get('new_status')
                    timestamp = entry.get('timestamp', '')[:19]
                    print(f"    {i}. {old} → {new} ({timestamp})")
    else:
        print(f"✗ Update failed: {result.get('error', 'Unknown error')}")


def demo_filter_by_status():
    """Demonstrate filtering jobs by status"""
    print_section("Demo 6: Filter Jobs by Status")
    
    storage_manager = JobStorageManager(storage_dir='demo_data')
    
    statuses = ['Applied', 'Interview', 'Offer', 'Pending', 'Rejected']
    
    print(f"\nFiltering jobs by status:")
    for status in statuses:
        jobs = storage_manager.get_jobs_by_status(status)
        print(f"\n  {status}: {len(jobs)} job(s)")
        for job in jobs:
            print(f"    - {job['title']} at {job['company']}")


def cleanup_demo_files():
    """Clean up demo files"""
    import shutil
    
    if os.path.exists('demo_data'):
        shutil.rmtree('demo_data')
        print("\n✓ Cleaned up demo data")


def main():
    """Run all demos"""
    print("=" * 70)
    print(" Excel Upload Module - Interactive Demo")
    print(" Task 7.3: Excel Upload for Status Tracking")
    print("=" * 70)
    
    if not OPENPYXL_AVAILABLE:
        print("\n⚠ Warning: openpyxl not installed")
        print("Some features will be limited")
        print("Install with: pip install openpyxl")
        print("\nContinuing with available features...\n")
    
    sample_file = 'sample_job_applications.xlsx'
    
    try:
        # Demo 1: Create sample Excel file
        if OPENPYXL_AVAILABLE:
            create_sample_excel(sample_file)
            
            # Demo 2: Parse Excel
            parse_result = demo_parse_excel(sample_file)
            
            if parse_result:
                # Demo 3: Validate against storage
                demo_validate_against_storage(parse_result)
                
                # Demo 4: Status updates
                demo_status_updates(parse_result)
        
        # Demo 5: Status summary
        demo_status_summary()
        
        # Demo 6: Single status update
        demo_single_status_update()
        
        # Demo 7: Filter by status
        demo_filter_by_status()
        
        print_section("Demo Complete")
        print("\nAll demonstrations completed successfully!")
        print("\nKey Features Demonstrated:")
        print("  ✓ Excel file parsing and validation")
        print("  ✓ Data integrity checks")
        print("  ✓ Status validation (Applied, Interview, Offer, Rejected, Pending)")
        print("  ✓ Validation against stored jobs")
        print("  ✓ Status update extraction")
        print("  ✓ Batch status updates")
        print("  ✓ Single job status updates")
        print("  ✓ Status history tracking")
        print("  ✓ Filtering by status")
        print("  ✓ Status summary statistics")
        
    except Exception as e:
        print(f"\n✗ Demo failed with error: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Cleanup
        if os.path.exists(sample_file):
            os.remove(sample_file)
            print(f"\n✓ Removed sample file: {sample_file}")
        
        cleanup_demo_files()


if __name__ == '__main__':
    main()
